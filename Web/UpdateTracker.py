from openpyxl.styles import Font, PatternFill


class UpdateTracker:

    def __init__(self,trackersht):
        self.trackersht = trackersht

    def update_po_consigmentstatus(self,pono,status):
        for i in range(2, 10000):
            if str(self.trackersht.cell(row=i, column=2).value).strip() == pono:
                self.trackersht.cell(row=i, column=9).value=status
                self.trackersht.cell(row=i, column=10).value = status

    def update_scheduleddate(self,pono,scheduleddate,consignmentno):
        for i in range(2, 10000):
            if str(self.trackersht.cell(row=i, column=2).value).strip() == pono:
                self.trackersht.cell(row=i, column=12).value=scheduleddate
                self.trackersht.cell(row=i, column=10).value=consignmentno


    def update_dateandtime(self,pono,dateandtime):
        for i in range(2, 10000):
            if str(self.trackersht.cell(row=i, column=2).value).strip() == pono:
                self.trackersht.cell(row=i, column=11).value=dateandtime

    def update_consignmentstatus(self, pono, status):
        for i in range(2, 10000):
            if str(self.trackersht.cell(row=i, column=2).value).strip() == pono:
                if len(status)<15:
                    self.trackersht.cell(row=i, column=10).value = status
                    green_cell = self.trackersht.cell(row=i, column=10)
                    green_cell.font = Font(color="008000", bold=False)
                else:
                    self.trackersht.cell(row=i, column=10).value = status
                    red_cell = self.trackersht.cell(row=i, column=10)
                    red_cell.font = Font(color="FF0000", bold=False)

    def update_postatus(self,pono,status):
        for i in range(2, 10000):
            if str(self.trackersht.cell(row=i, column=2).value).strip() == pono:
                if status == "Approved":
                    self.trackersht.cell(row=i, column=9).value = status
                    green_cell = self.trackersht.cell(row=i, column=9)
                    green_cell.font = Font(color="008000", bold=False)
                else:
                    self.trackersht.cell(row=i, column=9).value = status
                    red_cell = self.trackersht.cell(row=i, column=9)
                    red_cell.font = Font(color="FF0000", bold=False)

    def get_availableqty(self,ponum,product):
        result=""
        for i in range(2, 10000):
            if str(self.trackersht.cell(row=i, column=2).value).strip() == ponum and str(self.trackersht.cell(row=i, column=5).value).strip() == product:
                result = self.trackersht.cell(row=i, column=8).value
        if result=="":
            return 0
        elif result is None:
            return 0
        else:
            return result

    def update_slottedqty(self,ponum,product,slottedqty):
        for i in range(2, 10000):
            if str(self.trackersht.cell(row=i, column=2).value).strip() == ponum and str(self.trackersht.cell(row=i, column=5).value).strip() == product:
                self.trackersht.cell(row=i, column=13).value=slottedqty

    def error(self,ponum,product):
        for i in range(2,10000):
            if str(self.trackersht.cell(row=i, column=2).value).strip() == ponum and str(self.trackersht.cell(row=i, column=5).value).strip() == product:
                self.trackersht.cell(row=i, column=13).value = "quantityerror"

    def noslots(self):
        for i in range(2, 10000):
            if str(self.trackersht.cell(row=i, column=1).value).strip() != "None" and str(self.trackersht.cell(row=i, column=1).value).strip() != "":
                if (str(self.trackersht.cell(row=i, column=9).value))!="Approved":
                    self.trackersht.cell(row=i, column=13).value=""








