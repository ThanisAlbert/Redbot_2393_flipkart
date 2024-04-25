from tkinter import messagebox
import openpyxl
from Log.Log import Log
from Web.UpdateTracker import UpdateTracker
from Web.WebNavigate import WebNavigate

class Tracker(Log):

    def __init__(self,tracker,username,password,label):
        self.tracker = tracker
        self.username = username
        self.password = password
        self.label = label
        self.log = self.getLogger()

    def read(self):

        print(self.tracker)
        tracker_wk = openpyxl.load_workbook(self.tracker, read_only=False, keep_vba=True)
        tracker_sht = tracker_wk['Sheet1']

        maxrow = 0
        for i in range(2, 10000):
            if str(tracker_sht.cell(row=i, column=1).value).strip() != "None" and str(tracker_sht.cell(row=i, column=1).value).strip() != "":
                tracker_sht.cell(row=i, column=9).value=""
                tracker_sht.cell(row=i, column=10).value=""
                tracker_sht.cell(row=i, column=11).value=""
                tracker_sht.cell(row=i, column=12).value=""
                tracker_sht.cell(row=i, column=13).value=""
                tracker_sht.cell(row=i, column=14).value=""
                tracker_sht.cell(row=i, column=15).value= ""
                maxrow = maxrow + 1

        tracker_wk.save(self.tracker)
        tracker_wk.close()

        self.label.configure(text=str("Logging in "))
        self.label.update()

        flipkart = WebNavigate(username=self.username,password=self.password)
        flipkart.login()

        ponos=set()
        for i in range(2, maxrow + 2):
            ponos.add(tracker_sht.cell(row=i,column=2).value)

        i=1
        for pono in ponos:

            self.label.configure(text=str(f"Creating  {i} of {len(ponos)}"))
            self.label.update()

            result = self.validpo(tracker_sht,pono)

            if result==1:
                try:
                    flipkart.draftconsignment(ponum=pono, tracker_sht=tracker_sht)
                except Exception as e:
                    self.log.info("login_page_load_error")
                    self.log.info(e)
                    updatetracker = UpdateTracker(trackersht=tracker_sht)
                    updatetracker.update_postatus(pono, "technicalissue")

            i=i+1

        tracker_wk.save(self.tracker)
        tracker_wk.close()
        self.label.configure(text=str(f"Completed"))
        self.label.update()
        messagebox.showinfo('Info', 'Consignment Draft Completed!')

    def validpo(self,tracker_sht,pono):
        count=0
        for i in range(2, 10000):
            if str(tracker_sht.cell(row=i, column=1).value).strip() != "None" and str(tracker_sht.cell(row=i, column=1).value).strip() != "":
                if str(tracker_sht.cell(row=i, column=2).value).strip()==pono:
                    count=count+1
        if count>48:
            updatetracker = UpdateTracker(trackersht=tracker_sht)
            updatetracker.update_postatus(pono, "Please_do_manual_for_this_po")
            return 0
        else:
            return 1


    def createconsignment(self):

        tracker_wk = openpyxl.load_workbook(self.tracker)
        tracker_sht = tracker_wk['Sheet1']
        maxrow = 0  # validation required
        for i in range(2, 10000):
            if str(tracker_sht.cell(row=i, column=1).value).strip() != "None" and str(tracker_sht.cell(row=i, column=1).value).strip() != "":
                tracker_sht.cell(row=i, column=15).value=""
                maxrow = maxrow + 1

        tracker_wk.save(self.tracker)
        self.label.configure(text=str("Loggin in "))
        self.label.update()

        flipkart = WebNavigate(username=self.username, password=self.password)
        flipkart.login()

        self.label.configure(text=str(f"Creating Consignments.."))
        self.label.update()

        consignmentsset=set()
        for i in range(2, maxrow + 2):
            consignmentsval = str(tracker_sht.cell(row=i, column=10).value).strip()
            selectdateval = str(tracker_sht.cell(row=i, column=14).value).strip()
            remarks = str(tracker_sht.cell(row=i, column=15).value).strip()
            if len(consignmentsval) >= 10 and len(consignmentsval) <= 12 and selectdateval != "None" and selectdateval != "" and remarks == "":
                consignmentsset.add(consignmentsval)
        totalconsignments = len(consignmentsset)

        count = 1
        for i in range(2, maxrow + 2):

            consignmentsval = str(tracker_sht.cell(row=i, column=10).value).strip()
            selectdateval = str(tracker_sht.cell(row=i, column=14).value).strip()
            remarks = str(tracker_sht.cell(row=i, column=15).value).strip()

            if len(consignmentsval)>=10 and len(consignmentsval)<=12 and selectdateval!="None" and selectdateval!="" and remarks=="":

                self.label.configure(text=str(f"Consignments {count} of {totalconsignments}"))
                self.label.update()

                try:
                    fcnno = flipkart.createconsignment(consignmentsval, selectdateval)
                except Exception as e:
                    self.log.info(e)
                    fcnno="fcnerror"

                for i in range(2, maxrow + 2):
                    track_consignmentval = str(tracker_sht.cell(row=i, column=10).value).strip()
                    if consignmentsval==track_consignmentval:
                        tracker_sht.cell(row=i, column=11).value = fcnno
                        tracker_sht.cell(row=i, column=15).value="Done"

                count = count + 1

        tracker_wk.save(self.tracker)
        self.label.configure(text=str(f"Completed"))
        self.label.update()
        messagebox.showinfo('Info', 'Consignment Creation Completed!')






















        #flipkart = WebNavigate(username=self.username, password=self.password)
        #flipkart.login()










